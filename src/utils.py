from ics import Calendar
import csv
#import streamlit as st
import pandas as pd
from datetime import datetime
import locale
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import glob

def preprocess(file_object, output_buffer):
    """
    Adjust the formatting of LOCATION lines so that the next two lines are indented
    and replace occurrences of "Ny Stadion" with Ny Stadion.
    """
    lines = file_object.read().decode('utf-8').split('\n')

    # Combine all lines into a single string for easier processing
    content = ''.join(lines)

    # Replace "Ny Stadion" with Ny Stadion
    content = content.replace('"Ny Stadion"', 'Ny Stadion')

    # Split the content back into lines
    lines = content.splitlines()

    processed_lines = []
    i = 0
    while i < len(lines):
        line = lines[i]
        processed_lines.append(line + "\n")

        # Check if this line is a LOCATION line
        if line.startswith("LOCATION:"):
            # Indent the next two lines if they exist
            if i + 1 < len(lines) and not lines[i + 1].startswith(" "):
                processed_lines.append(" " + lines[i + 1].strip() + "\n")
                i += 1
            if i + 1 < len(lines) and not lines[i + 1].startswith(" "):
                processed_lines.append(" " + lines[i + 1].strip() + "\n")
                i += 1
        i += 1

    # Write the processed lines to the output buffer
    output_buffer.write(''.join(processed_lines).encode('utf-8'))
    output_buffer.seek(0)

def parse_ics_to_csv(ics_files, csv_file):
    """Parse multiple ICS files and export events to CSV."""
    all_events = []

    # Handle single file or list of files
    if not isinstance(ics_files, list):
        ics_files = [ics_files]

    # Collect events from all files
    for ics_file in ics_files:
        # Read content directly from BytesIO object
        content = ics_file.read().decode('utf-8')
        ics_file.seek(0)  # Reset buffer position for potential reuse
        calendar = Calendar(content)
        all_events.extend(calendar.events)

    # Sort events by date and time
    all_events.sort(key=lambda x: x.begin)

    # Write to CSV
    with open(csv_file, mode='w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['Description']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()

        for event in all_events:
            writer.writerow({
                'Description': event.description or ''
            })
            
def parse_entry(content):
    try:
        # Regex patterns to extract fields
        mnr_pat = r"Kampnr (\d+)"
        dt_pat = r"(\d{2}-\d{2}-\d{4}) kl\. (\d{2}:\d{2})"

        # Extract data using regex
        mnr_match = re.search(mnr_pat, content)
        dt_match = re.search(dt_pat, content)

        # Assign default values if matches are missing
        mnr = mnr_match.group(1) if mnr_match else "Unknown"
        mnr = int(mnr)
        date, time = (
            (dt_match.group(1), dt_match.group(2))
            if dt_match
            else ("Unknown", "Unknown")
        )

        # Extract league name (Række) and Årgang
        div = content.split("\n")[0].strip()
        year = content.split("\n")[0].split(" ")[0]
        if not year.startswith("U"):
            year = "Senior"

        # Parse the date
        date_object = datetime.strptime(date, "%d-%m-%Y")

        # Get the week number
        weeknr = date_object.strftime("%V")
        weeknr = int(weeknr)

        # Get the day of the week
        wk_date = date_object.strftime("%A").capitalize()  # Full day name (e.g., Saturday)

        # Convert to danish
        wk_date = get_day_of_week_danish(wk_date)

        # Home and away teams
        home = content.split("\n")[3].split(" - ")[0]
        away = content.split("\n")[3].split(" - ")[1]

        # Region
        postnr = content.split("\n")[7].split(" ")[0]
        postnr = int(postnr)
        region = (
            "Øst" if postnr < 5000 else "Vest"
        )

        return [year, wk_date, date, weeknr, time, div, mnr, home, away, region]
    except Exception as e:
        print(f"Error parsing entry: {content}\n{e}")
        return None

def get_day_of_week_danish(day_of_week):
    convert = {
        "Monday": "Mandag",
        "Tuesday": "Tirsdag",
        "Wednesday": "Onsdag",
        "Thursday": "Torsdag",
        "Friday": "Fredag",
        "Saturday": "Lørdag",
        "Sunday": "Søndag"
    }

    return convert.get(day_of_week, day_of_week)

def mk_df():
    df = pd.DataFrame(columns=["Årgang", "Dag", "Dato", "Uge", "Tidspunkt", "Række", "Kampnr", "Hjem", "Ude", "Region"])

    return df

def fill_df(df, input_path):
    with open(input_path, 'r', encoding='utf-8') as f:
        raw_data = f.read()

    # Split the data into entries by matching content between double quotes
    entries = re.findall(r'"(.*?)"', raw_data, re.DOTALL)
    fixed_entries = [item.replace("*** IKKE TIDS FASTSAT ****\n\n", "") for item in entries]
    pattern = r'\nInkl\. straffe: \d+-\d+'
    cleaned_entries = [re.sub(pattern, "", item) for item in fixed_entries]
    for entry in cleaned_entries:
        df.loc[len(df)] = parse_entry(entry)

    df["Scout"] = ""

    return df

def to_excel(df, buffer):
    """
    Converts a DataFrame to a styled Excel file and writes it to a BytesIO buffer.
    """
    # Replace problematic characters
    df = df.map(lambda x: x.replace("��", "ø") if isinstance(x, str) else x)

    # Process and sort the DataFrame
    df['Dato'] = pd.to_datetime(df['Dato'], format='%d-%m-%Y').dt.date
    df.sort_values(by=["Dato", "Tidspunkt"], ascending=[True, True], inplace=True)
    
    # Save DataFrame to an Excel buffer
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    # Load the workbook from the buffer
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active

    # Adjust column widths
    ws.column_dimensions['C'].width = 11.5
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['F'].width = 33
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 22

    # Define the blue fill for every second row
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Apply blue background to every second row starting from row 2
    for row in range(2, len(df) + 2, 2):  # Start from row 2 (header is row 1)
        for cell in ws[row]:
            cell.fill = blue_fill

    # Save the workbook back to the buffer
    buffer.seek(0)
    wb.save(buffer)
    buffer.seek(0)
