from ics import Calendar
import csv
import streamlit as st
import pandas as pd
from datetime import datetime
import locale
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def preprocess(input_path, output_path):
    """
    Adjust the formatting of LOCATION lines so that the next two lines are indented.
    """
    with open(input_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    processed_lines = []
    i = 0
    while i < len(lines):
        line = lines[i]
        processed_lines.append(line)

        # Check if this line is a LOCATION line
        if line.startswith("LOCATION:"):
            # Indent the next two lines if they exist
            if i + 1 < len(lines) and not lines[i + 1].startswith(" "):
                processed_lines.append(" " + lines[i + 1].strip() + "\n")
                i += 1
            if i + 1 < len(lines) and not lines[i + 1].startswith(" "):
                processed_lines.append(" " + lines[i + 1].strip() + "\n")
                i += 1
        i += 1

    # Write the processed lines to a new file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.writelines(processed_lines)

def parse_ics_to_csv(ics_file, csv_file):
    """Parse ICS file and export events to CSV."""
    with open(ics_file, 'r', encoding='utf-8') as f:
        content = f.read()
        calendar = Calendar(content)

    # Open CSV file for writing
    with open(csv_file, mode='w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['Description']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()

        for event in calendar.events:
            writer.writerow({
                'Description': event.description or ''
            })
            

def parse_entry(content):
    try:
        # Regex patterns to extract fields
        kampnr_pattern = r"Kampnr (\d+)"
        date_time_pattern = r"(\d{2}-\d{2}-\d{4}) kl\. (\d{2}:\d{2})"

        # Extract data using regex
        kampnr_match = re.search(kampnr_pattern, content)
        date_time_match = re.search(date_time_pattern, content)

        # Assign default values if matches are missing
        kampnr = kampnr_match.group(1) if kampnr_match else "Unknown"
        kampnr = int(kampnr)
        dato, tidspunkt = (
            (date_time_match.group(1), date_time_match.group(2))
            if date_time_match
            else ("Unknown", "Unknown")
        )

        # Extract league name (Række) and Årgang
        række = content.split("\n")[0].strip()
        årgang = content.split("\n")[0].split(" ")[0]  # Hardcoded based on the description

        # Find the day of the week
        locale.setlocale(locale.LC_TIME, "danish")
        date_string = "10-11-2024"  # Format: DD-MM-YYYY

        # Parse the date
        date_object = datetime.strptime(date_string, "%d-%m-%Y")

        # Get the day of the week
        dag = date_object.strftime("%A").capitalize()  # Full day name (e.g., Saturday)

        # Home and away teams
        hjem = content.split("\n")[3].split(" - ")[0]
        ude = content.split("\n")[3].split(" - ")[1]

        # Region
        postnr = content.split("\n")[7].split(" ")[0]
        postnr = int(postnr)
        region = (
            "Øst" if postnr < 5000 else "Vest"
        )

        return [årgang, dag, dato, tidspunkt, række, kampnr, hjem, ude, region]
    except Exception as e:
        print(f"Error parsing entry: {content}\n{e}")
        return None
    
def mk_df():
    df = pd.DataFrame(columns=["Årgang", "Dag", "Dato", "Tidspunkt", "Række", "Kampnr", "Hjem", "Ude", "Region"])

    return df

def fill_df(df, input_path):
    with open(input_path, 'r', encoding='utf-8') as f:
        raw_data = f.read()

    # Split the data into entries by matching content between double quotes
    entries = re.findall(r'"(.*?)"', raw_data, re.DOTALL)
    for entry in entries:
        df.loc[len(df)] = parse_entry(entry)

    return df

def to_excel(df, output_path):
    df = df.map(lambda x: x.replace("��", "ø") if isinstance(x, str) else x)
    df['Dato'] = pd.to_datetime(df['Dato'], format='%d-%m-%Y').dt.date
    df.sort_values(by=["Dato", "Tidspunkt"], ascending=[True, True], inplace=True)
    df.to_excel(f"{output_path}", index=False)

    # Make columns wider
    wb = load_workbook(output_path)
    ws = wb.active

    ws.column_dimensions['C'].width = 11.5
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 22

    # Define the blue fill for every second row
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Apply blue background to every second row starting from row 1 (Excel rows start at 1)
    for row in range(1, len(df) + 2, 2):  # Start from row 1 (Excel row 1 corresponds to pandas header)
        for cell in ws[row]:
            cell.fill = blue_fill

    wb.save(output_path)